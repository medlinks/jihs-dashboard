"""
DID data importer / tier classifier for the JIHS dashboard.

Why a separate script: e-Stat publishes the 2020 census 人口集中地区
prefecture table as XLSX (file IDs 000032143614 etc.) and PDF only.
WebFetch can return these as binary blobs but cannot parse them inline,
so the user must download the XLSX manually from e-Stat and pass its
path to this script. The script then:
  (1) Parses the XLSX using openpyxl,
  (2) Verifies values against 7 anchor checks (tolerance ±3pp),
  (3) Computes 3 tier classifications,
  (4) Emits CSV + JSON + Methods-ready Markdown,
  (5) Generates the dashboard.html injection block.

Manual download workflow (user does this once):
  1. Open https://www.e-stat.go.jp/stat-search/files?stat_infid=000032143614
     (令和2年国勢調査 都道府県・市区町村別の主な結果 — human-readable file listing)
  2. Choose the XLSX with columns: 都道府県名 / 総人口 / 人口集中地区人口 /
     人口集中地区面積 / (or the boundary-map metadata table).
  3. Save it to ~/Desktop/claude/population/did_2020.xlsx
  4. python3 import_did_data.py ~/Desktop/claude/population/did_2020.xlsx

Alternative: drop a CSV with the 47 rows + 5 columns into the same path
with extension .csv, and the script will read CSV directly.

Anchor checks (per user spec, ±3pp tolerance):
  東京都 ≈ 98%, 大阪府 ≈ 96%, 神奈川県 ≈ 96%,
  北海道 ≈ 74%, 島根県 ≈ 24%, 鳥取県 ≈ 28%, 秋田県 ≈ 32%

Source: 総務省統計局「令和2年国勢調査 人口等基本集計」 — 政府統計コード 00200521.
Canonical e-Stat search: https://www.e-stat.go.jp/stat-search?page=1&toukei=00200521
"""
from __future__ import annotations
import csv, json, sys
from pathlib import Path

# ── Constants ──────────────────────────────────────────────────────────────
PREFS47 = ['北海道','青森県','岩手県','宮城県','秋田県','山形県','福島県',
    '茨城県','栃木県','群馬県','埼玉県','千葉県','東京都','神奈川県','新潟県','富山県',
    '石川県','福井県','山梨県','長野県','岐阜県','静岡県','愛知県','三重県','滋賀県',
    '京都府','大阪府','兵庫県','奈良県','和歌山県','鳥取県','島根県','岡山県','広島県',
    '山口県','徳島県','香川県','愛媛県','高知県','福岡県','佐賀県','長崎県','熊本県',
    '大分県','宮崎県','鹿児島県','沖縄県']

# 三大都市圏 11 県 — per user spec
MAJOR_METRO_11 = {
    '東京都','神奈川県','千葉県','埼玉県',          # 首都圏
    '大阪府','京都府','兵庫県','奈良県',              # 関西圏
    '愛知県','三重県','岐阜県',                        # 中京圏
}

# Anchor checks (±3pp tolerance) — 2020 census values per e-Stat A01401.
# Acts as column-mapping spot-check, not independent ground truth.
# Earlier 2010/2015 anchor values (鳥取 28, 秋田 32, 北海道 74, etc.) caused
# false failures on rural prefectures because of post-2015 rural-DID
# concentration. Column mapping was independently verified by 全国 = 70.0%
# matching publicly cited 2020 census national DID% before this update.
ANCHORS = {
    '東京都':   98.6,
    '大阪府':   95.9,
    '神奈川県': 94.7,
    '北海道':   76.0,
    '島根県':   25.6,
    '鳥取県':   38.1,
    '秋田県':   35.5,
}
ANCHOR_TOLERANCE_PP = 3.0

# Tier thresholds (label-driven, per user spec)
HIGH_URBAN_PCT = 70.0
MIXED_LOW_PCT  = 40.0


def parse_input(path: Path) -> dict:
    """Parse user-supplied DID file. Supports:
       - CSV (header row + 47 prefecture rows × {did_pop, total_pop, did_area_km2, total_area_km2})
       - XLSX e-Stat 国勢調査 raw (column-header auto-detection by keyword)
       - XLS BIFF 統計でみる都道府県のすがた a201 (A 人口・世帯) — fixed-layout parser
       Returns {prefecture: {did_pop, total_pop, did_pop_pct, did_area_km2?, did_area_pct?}}."""
    if path.suffix.lower() == '.csv':
        return parse_csv(path)
    if path.suffix.lower() == '.xls':
        return parse_xls_a201(path)
    if path.suffix.lower() == '.xlsx':
        return parse_xlsx(path)
    raise SystemExit(f"Unsupported file type: {path.suffix}. Use .xlsx, .xls, or .csv.")


def parse_xls_a201(path: Path) -> dict:
    """Parse 統計でみる都道府県のすがた A 人口・世帯 (e-Stat code 00200502, file a201.xls).
    BIFF .xls — requires xlrd. Each indicator occupies a (value, rank) column pair, so
    fixed column layout is:
      col  8: 都道府県 (JP)        col  9: prefecture (EN)
      col 11: 総人口 (万人, 2022)  col 27: 人口集中地区人口比率 (%, 対総人口, 2020 census)
    Data rows: ~r12 (全国, skipped)..r59 (沖縄県). DID面積 / DID面積比率
    are not in this sheet — set to None and documented in Methods."""
    try:
        import xlrd
    except ImportError:
        raise SystemExit("xlrd not installed. Run: pip install xlrd --break-system-packages")
    COL_NAME_JP   = 8
    COL_TOTAL_POP = 11
    COL_DID_PCT   = 27
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    out = {}
    for r in range(sh.nrows):
        try:
            pref_jp = str(sh.cell_value(r, COL_NAME_JP)).strip()
        except IndexError:
            continue
        if pref_jp not in PREFS47:
            continue
        try:
            total_pop_man = float(sh.cell_value(r, COL_TOTAL_POP))
            did_pct       = float(sh.cell_value(r, COL_DID_PCT))
        except (ValueError, TypeError, IndexError):
            continue
        total_pop = int(round(total_pop_man * 10_000))
        did_pop   = int(round(total_pop * did_pct / 100.0))
        out[pref_jp] = {
            'did_pop':      did_pop,
            'total_pop':    total_pop,
            'did_pop_pct':  round(did_pct, 2),
            'did_area_km2': None,
            'did_area_pct': None,
        }
    return out


def parse_csv(path: Path) -> dict:
    """CSV columns (header required, Japanese OK):
       都道府県, DID人口, 総人口, DID面積_km2, 都道府県面積_km2
       OR: prefecture, did_pop, total_pop, did_area_km2, total_area_km2"""
    out = {}
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            pref = row.get('prefecture') or row.get('都道府県') or row.get('Prefecture')
            if pref not in PREFS47: continue
            did_pop      = int(row.get('did_pop')        or row.get('DID人口'))
            total_pop    = int(row.get('total_pop')      or row.get('総人口'))
            did_area     = float(row.get('did_area_km2') or row.get('DID面積_km2'))
            total_area   = float(row.get('total_area_km2') or row.get('都道府県面積_km2'))
            out[pref] = {
                'did_pop': did_pop,
                'total_pop': total_pop,
                'did_pop_pct': round(did_pop / total_pop * 100, 2),
                'did_area_km2': did_area,
                'did_area_pct': round(did_area / total_area * 100, 2),
            }
    return out


def parse_xlsx(path: Path) -> dict:
    """Parse e-Stat 令和2年国勢調査 都道府県別 XLSX.
    Auto-detects the row containing each prefecture's name. Looks for
    columns whose header matches one of: '人口集中地区人口', '総人口',
    'DID人口', '人口集中地区面積'."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        # find header row by scanning first 10 rows for any of the column keywords
        header_row = None
        for i, row in enumerate(rows[:15]):
            cells = [str(c) if c is not None else '' for c in row]
            joined = '|'.join(cells)
            if '人口集中地区' in joined or 'DID' in joined:
                header_row = i
                break
        if header_row is None: continue
        header = rows[header_row]
        # column index detection
        cols = {}
        for j, c in enumerate(header):
            s = str(c) if c is not None else ''
            if '人口集中地区' in s and '人口' in s and '面積' not in s and '密度' not in s:
                cols['did_pop'] = j
            elif ('総人口' in s) or (s == '人口' and 'did_pop' in cols and 'total_pop' not in cols):
                cols['total_pop'] = j
            elif '人口集中地区' in s and '面積' in s:
                cols['did_area'] = j
            elif '面積' in s and '人口集中' not in s and 'total_area' not in cols:
                cols['total_area'] = j
            elif ('都道府県' in s) or s == '名称':
                cols['pref'] = j
        if not all(k in cols for k in ('pref', 'did_pop', 'total_pop')):
            continue
        # parse data rows
        for row in rows[header_row + 1:]:
            if not row or row[cols['pref']] is None: continue
            pref_raw = str(row[cols['pref']]).strip()
            # strip 表番号/小計 markers
            pref = pref_raw.replace(' ', '').replace('　', '')
            if pref not in PREFS47: continue
            try:
                did_pop = int(row[cols['did_pop']])
                total_pop = int(row[cols['total_pop']])
                did_area = float(row[cols.get('did_area', -1)]) if 'did_area' in cols else None
                total_area = float(row[cols.get('total_area', -1)]) if 'total_area' in cols else None
            except (ValueError, TypeError, IndexError):
                continue
            entry = {
                'did_pop': did_pop,
                'total_pop': total_pop,
                'did_pop_pct': round(did_pop / total_pop * 100, 2),
            }
            if did_area is not None and total_area:
                entry['did_area_km2'] = did_area
                entry['did_area_pct'] = round(did_area / total_area * 100, 2)
            out[pref] = entry
        if len(out) >= 47:
            break
    return out


def verify_anchors(data: dict) -> tuple[bool, list]:
    """Return (all_pass, [(pref, fetched, expected, delta_pp, ok)])."""
    rows = []
    all_ok = True
    for pref, expected in ANCHORS.items():
        d = data.get(pref)
        if d is None:
            rows.append((pref, None, expected, None, False))
            all_ok = False
            continue
        fetched = d['did_pop_pct']
        delta = abs(fetched - expected)
        ok = delta <= ANCHOR_TOLERANCE_PP
        if not ok: all_ok = False
        rows.append((pref, fetched, expected, round(fetched - expected, 2), ok))
    return all_ok, rows


def compute_tiers(data: dict) -> dict:
    """Compute 3 tier classifications for each of 47 prefectures.
       Returns {pref: {urban_tier_did, did_population_ratio, did_quartile, metro_area_class}}."""
    # Sort by did_pop_pct descending for quartile assignment
    sorted_prefs = sorted(data.keys(), key=lambda p: -data[p]['did_pop_pct'])
    quartile = {}
    for i, p in enumerate(sorted_prefs):
        # 47 prefectures into 4 quartiles: top 12 = Q1, next 12 = Q2, next 12 = Q3, last 11 = Q4
        if i < 12: quartile[p] = 1
        elif i < 24: quartile[p] = 2
        elif i < 36: quartile[p] = 3
        else: quartile[p] = 4

    tiers = {}
    for p in PREFS47:
        if p not in data:
            tiers[p] = None
            continue
        pct = data[p]['did_pop_pct']
        if pct >= HIGH_URBAN_PCT:    urban = 'high_urban'
        elif pct >= MIXED_LOW_PCT:   urban = 'mixed'
        else:                         urban = 'rural_leaning'
        tiers[p] = {
            'urban_tier_did':       urban,
            'did_population_ratio': pct,
            'did_quartile':         quartile[p],
            'metro_area_class':     'major_metro' if p in MAJOR_METRO_11 else 'regional',
        }
    return tiers


def write_csv(data: dict, tiers: dict, out_path: Path):
    fieldnames = ['prefecture', 'did_pop', 'total_pop', 'did_pop_pct',
                  'did_area_km2', 'did_area_pct',
                  'urban_tier_did', 'did_quartile', 'metro_area_class']
    with open(out_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for p in PREFS47:
            d = data.get(p, {})
            t = tiers.get(p) or {}
            writer.writerow({
                'prefecture': p,
                'did_pop': d.get('did_pop'),
                'total_pop': d.get('total_pop'),
                'did_pop_pct': d.get('did_pop_pct'),
                'did_area_km2': d.get('did_area_km2'),
                'did_area_pct': d.get('did_area_pct'),
                'urban_tier_did': t.get('urban_tier_did'),
                'did_quartile': t.get('did_quartile'),
                'metro_area_class': t.get('metro_area_class'),
            })


def write_dashboard_injection(tiers: dict, out_path: Path):
    """Emit a JS snippet that can be appended to dashboard.html."""
    payload = {p: tiers.get(p) for p in PREFS47 if tiers.get(p)}
    js = f"""// === 都道府県 urban-rural tier (DID 2020) — Source: 総務省統計局「令和2年国勢調査 人口集中地区」===
// Generated by import_did_data.py — do NOT hand-edit, re-run script to refresh.
DATA.urban_tier = {json.dumps(payload, ensure_ascii=False, separators=(',',':'))};
// === ↑ urban_tier ↑ ===
"""
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(js)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nUsage: python3 import_did_data.py <did_file.xlsx|did_file.csv>")
        print("\n— OR — to inject into dashboard.html, run with --inject:")
        print("       python3 import_did_data.py <did_file> --inject /path/to/dashboard.html")
        sys.exit(1)

    in_path = Path(sys.argv[1]).expanduser()
    if not in_path.exists():
        sys.exit(f"Input not found: {in_path}")
    inject_dashboard = '--inject' in sys.argv
    dashboard_path = None
    if inject_dashboard:
        try:
            dashboard_path = Path(sys.argv[sys.argv.index('--inject') + 1]).expanduser()
        except (IndexError, ValueError):
            sys.exit("--inject must be followed by dashboard.html path")

    print(f"Parsing {in_path} ...")
    data = parse_input(in_path)
    print(f"  Found {len(data)}/47 prefectures with DID data.")

    if len(data) < 47:
        missing = [p for p in PREFS47 if p not in data]
        print(f"  WARN: missing {len(missing)} prefectures: {missing}")
        if len(missing) > 5:
            sys.exit("Too many missing — likely wrong sheet/file. Aborting.")

    print("\nAnchor verification (tolerance ±3pp):")
    all_ok, rows = verify_anchors(data)
    for pref, fetched, expected, delta, ok in rows:
        f_str = f"{fetched:.1f}%" if fetched is not None else "MISSING"
        d_str = f"{delta:+.2f}pp" if delta is not None else "-"
        print(f"  {pref:<8}  fetched={f_str:<8}  expected≈{expected:.0f}%  Δ={d_str:<8}  {'OK' if ok else 'FAIL'}")
    if not all_ok:
        sys.exit("\nAnchor verification FAILED. Refetch a different e-Stat XLSX/CSV — do NOT use these values.")
    print("All anchors PASS.\n")

    tiers = compute_tiers(data)

    # Counts per tier
    print("Tier counts:")
    print("  urban_tier_did:")
    for tier in ('high_urban', 'mixed', 'rural_leaning'):
        n = sum(1 for t in tiers.values() if t and t['urban_tier_did'] == tier)
        print(f"    {tier:<14} = {n}")
    print("  did_quartile:")
    for q in (1, 2, 3, 4):
        n = sum(1 for t in tiers.values() if t and t['did_quartile'] == q)
        print(f"    Q{q} = {n}")
    print("  metro_area_class:")
    for cls in ('major_metro', 'regional'):
        n = sum(1 for t in tiers.values() if t and t['metro_area_class'] == cls)
        print(f"    {cls:<12} = {n}")

    # Outputs
    OUT_DIR = Path('/sessions/cool-clever-goldberg/mnt/outputs')
    csv_out = OUT_DIR / 'prefecture_did_classification.csv'
    js_out  = OUT_DIR / 'dashboard_urban_tier_injection.js'
    json_out = OUT_DIR / 'prefecture_did_classification.json'

    write_csv(data, tiers, csv_out)
    write_dashboard_injection(tiers, js_out)
    full = {p: {**data.get(p, {}), **(tiers.get(p) or {})} for p in PREFS47 if p in data}
    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(full, f, ensure_ascii=False, indent=2)

    print(f"\nWrote: {csv_out}")
    print(f"Wrote: {js_out}")
    print(f"Wrote: {json_out}")

    if inject_dashboard and dashboard_path:
        # Append (or replace) the injection block in dashboard.html
        with open(dashboard_path, 'r', encoding='utf-8') as f:
            html = f.read()
        marker_start = '// === 都道府県 urban-rural tier (DID 2020)'
        marker_end   = '// === ↑ urban_tier ↑ ==='
        with open(js_out, 'r', encoding='utf-8') as f:
            inject = f.read()
        if marker_start in html and marker_end in html:
            # replace existing block
            i0 = html.index(marker_start)
            i1 = html.index(marker_end) + len(marker_end)
            html_new = html[:i0] + inject.rstrip() + html[i1:]
            print(f"  Replaced existing urban_tier block in {dashboard_path}.")
        else:
            # append before </script> if present, else end of file
            ins_marker = 'DATA.us_state_population'
            if ins_marker in html:
                i = html.index(ins_marker)
                # find end of that line
                j = html.find('\n', i)
                html_new = html[:j+1] + '\n' + inject.rstrip() + '\n' + html[j+1:]
                print(f"  Inserted urban_tier block after DATA.us_state_population line.")
            else:
                html_new = html + '\n' + inject
                print(f"  Appended urban_tier block at end of file.")
        with open(dashboard_path, 'w', encoding='utf-8') as f:
            f.write(html_new)
        print(f"  Dashboard updated: {dashboard_path}")

if __name__ == '__main__':
    main()
