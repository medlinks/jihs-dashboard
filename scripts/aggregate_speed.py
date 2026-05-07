"""
diseases/<疾病>/*.xlsx の「速報データ」シートを読み、
full_dashboard_data.json の speed_trends を再構築する。

collect_idwr.py が xlsx に書き込んだ最新の速報を、
dashboard が実際に読む speed_trends に反映する役割。
"""
import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DISEASES_DIR = ROOT / 'diseases'
DATA_FILE = ROOT / 'scripts' / 'full_dashboard_data.json'

# 「速報データ」シートのヘッダ列インデックス
COL_YEAR = 0
COL_WEEK = 1
COL_TOTAL = 4   # ['年','週','期間','作成日','総数', ...]


def aggregate():
    speed_trends = {}
    n_diseases = 0
    n_skipped = 0

    if not DISEASES_DIR.is_dir():
        raise SystemExit(f'diseases フォルダが見つかりません: {DISEASES_DIR}')

    for disease_dir in sorted(DISEASES_DIR.iterdir()):
        if not disease_dir.is_dir():
            continue
        xlsx_files = list(disease_dir.glob('*.xlsx'))
        if not xlsx_files:
            n_skipped += 1
            continue
        # 通常 1 ファイル
        xlsx = xlsx_files[0]
        disease_name = xlsx.stem

        try:
            wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
        except Exception as e:
            print(f'  SKIP {disease_name}: {e}')
            n_skipped += 1
            continue

        if '速報データ' not in wb.sheetnames:
            wb.close()
            n_skipped += 1
            continue

        ws = wb['速報データ']
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        entries = []
        seen = set()  # (year, week) で重複排除
        for row in rows:
            if not row or len(row) <= COL_TOTAL:
                continue
            y, w, total = row[COL_YEAR], row[COL_WEEK], row[COL_TOTAL]
            # 数値変換できない行（ヘッダ等）はスキップ
            try:
                year = int(y) if y is not None else None
                week = int(w) if w is not None else None
            except (ValueError, TypeError):
                continue
            if year is None or week is None:
                continue
            if not (2000 <= year <= 2100 and 1 <= week <= 53):
                continue
            key = (year, week)
            if key in seen:
                continue
            seen.add(key)
            try:
                tot = int(total) if total is not None else 0
            except (ValueError, TypeError):
                tot = 0
            entries.append({'year': year, 'week': week, 'total': tot})

        if entries:
            # 速報は直近のみ表示するが、全件保持し dashboard 側でフィルタ
            entries.sort(key=lambda r: (r['year'], r['week']))
            speed_trends[disease_name] = entries
            n_diseases += 1

    return speed_trends, n_diseases, n_skipped


def main():
    speed_trends, n, n_skip = aggregate()

    if not DATA_FILE.exists():
        raise SystemExit(f'full_dashboard_data.json が見つかりません: {DATA_FILE}')

    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    old_speed = data.get('speed_trends', {})
    n_changed = 0
    n_added = 0
    week_set_old = set()
    week_set_new = set()
    for d, lst in speed_trends.items():
        old = old_speed.get(d, [])
        if len(old) != len(lst) or any(
            o != n2 for o, n2 in zip(sorted(old, key=lambda r:(r['year'],r['week'])),
                                     sorted(lst, key=lambda r:(r['year'],r['week'])))
        ):
            n_changed += 1
        if d not in old_speed:
            n_added += 1
        for r in lst:
            week_set_new.add((r['year'], r['week']))
    for d, lst in old_speed.items():
        for r in lst:
            week_set_old.add((r['year'], r['week']))

    data['speed_trends'] = speed_trends

    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

    print(f'=== 速報集計 完了 ===')
    print(f'集計疾病数: {n}（スキップ {n_skip}）')
    print(f'変更あり: {n_changed} 件 / 新規 {n_added} 件')
    print(f'週セット 旧: {sorted(week_set_old)}')
    print(f'週セット 新: {sorted(week_set_new)}')
    diff = week_set_new - week_set_old
    if diff:
        print(f'★ 新たに含まれた週: {sorted(diff)}')


if __name__ == '__main__':
    main()
