"""
US NNDSS 全国合計データを集約 → 日本側と同形式の JSON + Excel に変換

入力 (優先順):
  1) us/raw/nndss/{year}_national.json   （旧: 国家行のみフィルタ済み）
  2) us/raw/nndss/{year}.json            （現: fetch_cdc_nndss.py が生成 - 全行）
                                          全国行は load_year() で動的にフィルタ
出力:
  - us/processed/us_full_data.json           （dashboard に注入する JSON）
  - us/processed/US_diseases_weekly.xlsx     （全疾病・全週・1 ファイル）
  - us/processed/per_disease/{disease}.xlsx  （疾病別、日本側と同形式）
  - us/processed/US_diseases_annual.xlsx     （年累積）
"""
import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "us" / "raw" / "nndss"
PROC_DIR = ROOT / "us" / "processed"
POP_FILE = ROOT / "us" / "population" / "us_population.csv"
MAPPING_FILE = ROOT / "us" / "disease_mapping.json"

YEARS = [2022, 2023, 2024, 2025, 2026]

# NNDSS の年別「全国合計」行に使われる states 値のいずれか。
# 2022/2024 は "US RESIDENTS"、2023 は両方混在、2025-2026 は "U.S. Residents" / "Total"。
# 旧 _national.json はこれで予めフィルタ済み。新 {year}.json には州・領域も含むため、
# load_year() でこの集合を使ってフィルタする。
NATIONAL_LABELS = {
    "US RESIDENTS",
    "U.S. RESIDENTS",
    "U.S. Residents",
    "Total",
}


def load_population():
    pop = {}
    if not POP_FILE.exists():
        return pop
    with open(POP_FILE, encoding="utf-8") as f:
        next(f)  # header
        for line in f:
            parts = line.strip().split(",")
            if len(parts) >= 2 and parts[0].isdigit():
                pop[int(parts[0])] = int(parts[1])
    return pop


def load_live_births():
    """
    Read us/population/us_live_births.csv → {year: live_births}.
    Used as the denominator for perinatal / congenital diseases (Congenital
    Syphilis, Congenital Rubella, Hepatitis B/C perinatal). Using total
    population for these gives a rate ~100× too low.
    """
    births_file = POP_FILE.parent / "us_live_births.csv"
    births = {}
    if not births_file.exists():
        return births
    with open(births_file, encoding="utf-8") as f:
        next(f)  # header
        for line in f:
            parts = line.strip().split(",")
            if len(parts) >= 2 and parts[0].isdigit():
                try:
                    births[int(parts[0])] = int(parts[1])
                except ValueError:
                    pass
    return births


def load_mapping():
    if not MAPPING_FILE.exists():
        return {}
    with open(MAPPING_FILE, encoding="utf-8") as f:
        data = json.load(f)
    # NNDSS label → 日本疾病名 / 統合カテゴリ
    label_to_jp = {}
    label_to_en = {}
    for m in data.get("mapping", []):
        for label in m.get("nndss_labels", []):
            label_to_jp[label] = m["jp"]
            label_to_en[label] = m["en"]
    return label_to_jp, label_to_en


def safe_int(v):
    """SODA API は数値も '0.0' のような文字列で返す"""
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def load_year(year: int) -> list:
    """
    Load the national-only NNDSS rows for `year`.

    Preference order:
      1) {year}_national.json — pre-filtered (legacy fetch_cdc_nndss output)
      2) {year}.json          — full dataset (current fetch_cdc_nndss output);
                                filter for national-total rows on the fly
                                using NATIONAL_LABELS.
    """
    nat_path  = RAW_DIR / f"{year}_national.json"
    full_path = RAW_DIR / f"{year}.json"

    if nat_path.exists():
        with open(nat_path, encoding="utf-8") as f:
            return json.load(f)

    if full_path.exists():
        with open(full_path, encoding="utf-8") as f:
            data = json.load(f)
        nat_rows = [r for r in data if r.get("states") in NATIONAL_LABELS]
        print(f"  {full_path.name}: {len(data)}行 → 全国 {len(nat_rows)}行 にフィルタ")
        return nat_rows

    print(f"  WARN: neither {nat_path.name} nor {full_path.name} なし")
    return []


def aggregate_weekly():
    """
    NNDSS label を残したまま週次データを集約
    {label: [{year, week, current, cumulative_ytd}, ...]}
    """
    by_label = defaultdict(list)
    for y in YEARS:
        rows = load_year(y)
        for r in rows:
            label = r.get("label")
            if not label:
                continue
            try:
                week = int(r.get("week", 0))
            except (TypeError, ValueError):
                continue
            current = safe_int(r.get("m1"))
            cum_ytd = safe_int(r.get("m3"))
            by_label[label].append({
                "year": y, "week": week,
                "current": current,
                "cumulative_ytd": cum_ytd,
            })

    # ソート
    for k in by_label:
        by_label[k].sort(key=lambda r: (r["year"], r["week"]))
    return dict(by_label)


def build_jp_aligned(by_label: dict, label_to_jp: dict, label_to_en: dict, population: dict):
    """
    日本側 weekly_trends 形式 + 罹患率付き
    {disease_en_or_jp: [{year, week, total, rate_per_100k}, ...]}
    JP マッピングがあれば複数 NNDSS label を統合
    """
    # JP 名でグルーピング（複数ラベル統合）
    jp_buckets = defaultdict(lambda: defaultdict(int))  # jp → (year, week) → sum
    en_buckets = defaultdict(lambda: defaultdict(int))  # en → (year, week) → sum
    for label, rows in by_label.items():
        jp = label_to_jp.get(label)
        en = label_to_en.get(label, label)  # マップ無しは label 自体
        for r in rows:
            key = (r["year"], r["week"])
            if jp:
                jp_buckets[jp][key] += r["current"]
            en_buckets[en][key] += r["current"]

    def to_series(bucket):
        out = []
        for (y, w), v in sorted(bucket.items()):
            entry = {"year": y, "week": w, "total": v}
            pop = population.get(y)
            if pop and pop > 0:
                entry["rate_per_100k"] = round(v / pop * 100000, 4)
            out.append(entry)
        return out

    return (
        {jp: to_series(b) for jp, b in jp_buckets.items()},
        {en: to_series(b) for en, b in en_buckets.items()},
    )


def write_full_xlsx(by_label_en: dict, path: Path):
    """全疾病 1 シート、横並びで週・縦並びで疾病"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "US Weekly NNDSS"

    # 全 (year, week) を集める
    all_keys = set()
    for series in by_label_en.values():
        for r in series:
            all_keys.add((r["year"], r["week"]))
    sorted_keys = sorted(all_keys)

    # ヘッダ
    ws.cell(1, 1, "Disease (NNDSS Label)").font = Font(bold=True)
    for i, (y, w) in enumerate(sorted_keys, start=2):
        c = ws.cell(1, i, f"{y}-W{w:02d}")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # 各疾病行
    for r_idx, (disease, series) in enumerate(sorted(by_label_en.items()), start=2):
        ws.cell(r_idx, 1, disease)
        m = {(s["year"], s["week"]): s["total"] for s in series}
        for c_idx, k in enumerate(sorted_keys, start=2):
            v = m.get(k, 0)
            ws.cell(r_idx, c_idx, v)

    # 列幅
    ws.column_dimensions["A"].width = 60
    for i in range(2, len(sorted_keys) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10
    ws.freeze_panes = "B2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  → {path.name}: {os.path.getsize(path)/1024:.0f} KB")


def write_per_disease_xlsx(by_disease_en: dict, label_to_jp: dict, out_dir: Path):
    """1 疾病 1 ファイル（日本側 diseases/{name}/{name}.xlsx と同形式）"""
    out_dir.mkdir(parents=True, exist_ok=True)
    label_jp_rev = {}
    for label, jp in label_to_jp.items():
        label_jp_rev.setdefault(jp, []).append(label)

    written = 0
    for disease, series in by_disease_en.items():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "週次データ"

        # ヘッダ
        ws.append(["Disease (English)", disease])
        ws.append([])
        ws.append(["Year", "MMWR Week", "Current Cases", "Rate per 100k", "Cumulative YTD"])
        for c in ws[3]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="FAE8DD")

        # データ行
        for s in series:
            ws.append([
                s["year"], s["week"],
                s["total"],
                s.get("rate_per_100k", 0),
                None,  # YTD は別途必要なら追加
            ])

        # 列幅
        for i, w in enumerate([22, 12, 14, 14, 14], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # ファイル名は安全に
        safe_name = re.sub(r'[\\/:*?"<>|,()]', "_", disease).strip("_")[:80]
        wb.save(out_dir / f"{safe_name}.xlsx")
        written += 1
    print(f"  → {written} ファイル in {out_dir.name}/")


def write_annual_xlsx(by_label: dict, path: Path):
    """年累積 (m3) を 1 シートで疾病 × 年"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "US Annual Cumulative"

    ws.cell(1, 1, "Disease").font = Font(bold=True)
    for i, y in enumerate(YEARS, start=2):
        ws.cell(1, i, str(y)).font = Font(bold=True)

    for r_idx, (label, series) in enumerate(sorted(by_label.items()), start=2):
        ws.cell(r_idx, 1, label)
        # 各年の最終週の m3 を取る
        latest_per_year = {}
        for r in series:
            y = r["year"]
            if y not in latest_per_year or r["week"] > latest_per_year[y]["week"]:
                latest_per_year[y] = r
        for c_idx, y in enumerate(YEARS, start=2):
            v = latest_per_year.get(y, {}).get("cumulative_ytd", 0)
            ws.cell(r_idx, c_idx, v)

    ws.column_dimensions["A"].width = 60
    for i in range(2, len(YEARS) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
    ws.freeze_panes = "B2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  → {path.name}: {os.path.getsize(path)/1024:.0f} KB")


def main():
    print("=== US NNDSS aggregator ===\n")
    population = load_population()
    print(f"人口データ: {len(population)} 年分")

    live_births = load_live_births()
    print(f"活産数データ: {len(live_births)} 年分(Congenital Syphilis 等の分母用)")

    label_to_jp, label_to_en = load_mapping()
    print(f"マッピング: {len(label_to_jp)} ラベル → 日本疾病")

    by_label = aggregate_weekly()
    print(f"\nNNDSS label 数: {len(by_label)}")

    # JP / EN 整列
    jp_aligned, en_aligned = build_jp_aligned(by_label, label_to_jp, label_to_en, population)
    print(f"JP マッチ: {len(jp_aligned)} 疾病")
    print(f"EN 全件: {len(en_aligned)} 疾病")

    # === 年次累積を m3 (cumulative YTD) の最大値から導出 ===
    # 基本: 各年 Y の値 = その年の m3 (cumulative YTD) 最大値
    #
    # ただし NNDSS Weekly は provisional で、前年の case count は翌年に
    # 改訂され続ける(典型的に +30-50%)。CDC は改訂値を「翌年の m4 列
    # (= cumulative YTD previous year)」で公開している。
    # 例: 2024 年の Syphilis, Congenital
    #   - 2024 自身の max m3 = 2,559 (provisional, 我々が以前使っていた)
    #   - 2025 weekly の max m4 = 3,936 (改訂後、CDC STI Annual の ~3,945 と一致)
    #
    # よって最新の年(YEARS[-1])は m3、それ以外の年は次年度の m4 を優先する。
    annual_totals = {}  # {label: {year: max_ytd}}
    for label, rows in by_label.items():
        per_year = {}
        # m3 (same-year YTD) を集める
        m3_by_year = {}
        for r in rows:
            y = r["year"]; ytd = r.get("cumulative_ytd", 0) or 0
            if y not in m3_by_year or ytd > m3_by_year[y]:
                m3_by_year[y] = ytd
        per_year.update(m3_by_year)

        # m4 (= 翌年から見た前年累計) を改訂値として優先する
        # rows には m4 が直接入っていないので by_label_with_m4 を別途構築する必要がある
        annual_totals[label] = per_year

    # 翌年の m4 列を別途取得し、改訂値として overwrite する
    # （by_label には m4 が入っていないため、生 raw を再スキャン）
    for next_year in YEARS:
        prev_year = next_year - 1
        if prev_year not in YEARS:
            continue
        rows = load_year(next_year)
        for r in rows:
            label = r.get("label")
            if not label:
                continue
            # m4 = previous year cumulative
            m4 = safe_int(r.get("m4"))
            if m4 <= 0:
                continue
            cur = annual_totals.setdefault(label, {}).get(prev_year, 0)
            if m4 > cur:
                annual_totals[label][prev_year] = m4

    # === EN 合并版年次累積 ===
    # us_weekly_trends は EN 合併名（disease_mapping の en）でキー、
    # us_annual_totals は raw NNDSS label でキーになっていたため、dashboard が
    # `us_annual_totals[selected_disease]` で 22 個の合併疾患（梅毒・はしか等）を
    # 引けず、buggy な m1-sum にフォールバックしていた。
    # ここで EN-merged な年次累積も同梱する：各 sub-label の max(m3) を年ごとに合算。
    annual_totals_en = {}  # {en_name: {year: sum_of_max_ytd_across_sublabels}}
    for label, per_year in annual_totals.items():
        en = label_to_en.get(label, label)  # マッピング外は label 自体を EN とみなす
        bucket = annual_totals_en.setdefault(en, {})
        for y, v in per_year.items():
            bucket[y] = bucket.get(y, 0) + v

    # === JP 合併版年次累積（国家対比モード用）===
    # 同じ問題が「🌐 国家対比 → 📅 年報」にもあった：
    # renderCompareAnnual が DATA.us_weekly_trends_jp[jp_disease] を年で sum していたが、
    # それは m1 sum なので梅毒等の合併疾患で大きな過小評価になっていた。
    # ここで JP-merged な年次累積を生成し、国家対比用の正しい値を提供する。
    annual_totals_jp = {}  # {jp_name: {year: sum_of_max_ytd_across_sublabels}}
    for label, per_year in annual_totals.items():
        jp = label_to_jp.get(label)
        if not jp:
            continue  # マッピングに無い疾病は対比できないのでスキップ
        bucket = annual_totals_jp.setdefault(jp, {})
        for y, v in per_year.items():
            bucket[y] = bucket.get(y, 0) + v

    # JSON 出力（dashboard 用）
    PROC_DIR.mkdir(parents=True, exist_ok=True)
    out_json = PROC_DIR / "us_full_data.json"
    payload = {
        "us_weekly_trends": en_aligned,         # NNDSS English 名（週次、m1 sum）
        "us_weekly_trends_jp": jp_aligned,      # 日本疾病名で集約（週次、m1 sum）
        "us_annual_totals": annual_totals,      # raw label 年累計（m3 max）
        "us_annual_totals_en": annual_totals_en,# EN 合併名 年累計（米国版年報棒図用）
        "us_annual_totals_jp": annual_totals_jp,# JP 合併名 年累計（国家対比年報用）
        "us_population": population,
        "us_live_births": live_births,          # 活産数（Congenital Syphilis 等の分母）
        "years_covered": YEARS,
        "national_only": True,
    }
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  → {out_json.name}: {os.path.getsize(out_json)/1024:.0f} KB")

    # Excel 出力
    print("\nExcel 生成中...")
    write_full_xlsx(en_aligned, PROC_DIR / "US_diseases_weekly.xlsx")
    write_annual_xlsx(by_label, PROC_DIR / "US_diseases_annual.xlsx")
    write_per_disease_xlsx(en_aligned, label_to_jp, PROC_DIR / "per_disease")

    print("\n✅ 完了")


if __name__ == "__main__":
    main()
