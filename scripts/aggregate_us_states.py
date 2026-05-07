"""
NNDSS 州別データを集約して dashboard 用 JSON に変換

入力: us/raw/nndss/states_{year}.json (複数年あれば結合)
出力:
  - us/processed/us_state_latest.json   {jp_disease: {state: {current, cumulative_ytd, year, week}}}
  - us/processed/us_state_latest_en.json (英語 NNDSS ラベル版)
  - us/processed/us_state_population.json
  - us/processed/state_per_disease/{disease}.xlsx
      Per-disease state × time matrix (rows = states, columns = year-week,
      cells = cumulative YTD). Single source of truth that any downstream
      tool — dashboard, ad-hoc analysis, alerting — can read.
"""
import json
import csv
import re
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "us" / "raw" / "nndss"
PROC_DIR = ROOT / "us" / "processed"
MAPPING = ROOT / "us" / "disease_mapping.json"
POP_FILE = ROOT / "us" / "population" / "us_state_population.csv"

# Sanitize disease names for filesystem use.
# Keep Unicode letters/digits (so 日本語/中文/한국어 disease names survive),
# but strip filesystem-unsafe punctuation: / \ : * ? " < > | etc.
_FS_UNSAFE = re.compile(r'[\\/:*?"<>|\x00-\x1f]+')
def _safe_filename(s: str) -> str:
    cleaned = _FS_UNSAFE.sub("_", s).strip(". _")
    # macOS / ext4 hard limit ~255 bytes; UTF-8 JP can be 3 bytes per char.
    # Encode and trim to keep us under 200 bytes safely.
    enc = cleaned.encode("utf-8")[:200]
    # In case we cut mid-codepoint, decode loosely
    return enc.decode("utf-8", errors="ignore") or "_unnamed"


def load_mapping_label_to_jp():
    with open(MAPPING, encoding="utf-8") as f:
        m = json.load(f)
    label_to_jp = {}
    for entry in m["mapping"]:
        for lab in entry.get("nndss_labels", []):
            label_to_jp[lab] = entry["jp"]
    return label_to_jp


def safe_int(v):
    if v is None or v == "" or v == "-":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def main():
    label_to_jp = load_mapping_label_to_jp()
    print(f"マッピング: {len(label_to_jp)} ラベル")

    # 集計: 2 つのキー方式で保存
    #   result_jp:  日本疾病名 → state → record（複数ラベル統合）
    #   result_en:  NNDSS 英語ラベル → state → record（個別ラベル）
    result = defaultdict(dict)
    result_en = defaultdict(dict)

    for path in sorted(RAW_DIR.glob("states_*.json")):
        with open(path, encoding="utf-8") as f:
            rows = json.load(f)
        print(f"  {path.name}: {len(rows)} 件")
        # ラベル統合用に (jp, state) → 各週レコード
        bucket = defaultdict(lambda: defaultdict(lambda: {"current":0, "cumulative_ytd":0, "year":None, "week":0}))
        for r in rows:
            label = r.get("label")
            jp = label_to_jp.get(label)
            if not jp:
                continue
            state = r.get("states")
            # Census divisions / regions / aggregates を除外
            CENSUS_DIVS = {
                "New England", "Middle Atlantic",
                "East North Central", "West North Central",
                "South Atlantic", "East South Central", "West South Central",
                "Mountain", "Pacific",
                # 古い大文字版
                "NEW ENGLAND", "MID. ATLANTIC", "MIDDLE ATLANTIC",
                "E.N. CENTRAL", "W.N. CENTRAL",
                "S. ATLANTIC", "E.S. CENTRAL", "W.S. CENTRAL",
                "MOUNTAIN", "PACIFIC",
            }
            EXCLUDE_KEYWORDS = ("Total", "Residents", "RESIDENTS", "Non-", "NON-", "Region")
            if (not state) or (state in CENSUS_DIVS) or any(k in state for k in EXCLUDE_KEYWORDS):
                continue
            try:
                year = int(r.get("year", 0))
                week = int(r.get("week", 0))
            except (TypeError, ValueError):
                continue
            cur = safe_int(r.get("m1"))
            ytd = safe_int(r.get("m3"))
            # 同 (jp, state, year, week) で複数ラベル合算
            key = (year, week)
            cell = bucket[(jp, state)][key]
            cell["current"] += cur
            cell["cumulative_ytd"] = max(cell["cumulative_ytd"], ytd)  # ラベル統合は max（同累積を二重カウントしない）
            cell["year"] = year
            cell["week"] = week
        # 各 (jp, state) について最新週を選ぶ
        for (jp, state), wks in bucket.items():
            keys = sorted(wks.keys())
            # ラベル統合は cumulative の sum がいい場合もあるが、ここでは「複数ラベル時は ytd を sum、current は sum」
            # 実際: 複数 NNDSS ラベル（例 Measles Imported + Indigenous）は別レコードで来る
            # bucket は (jp, state) → {(year,week): cell}
            # cell は cur=sum, ytd=max なので、Imported+Indigenous なら cur sum は OK だが ytd は max になる（不正確）
            # 修正: ytd も sum すべき → max にすると過小評価
            pass
        # === 英語ラベル直接版（全 119 疾病対応）===
        bucket_en = defaultdict(lambda: defaultdict(lambda: {"current":0, "cumulative_ytd":0, "year":None, "week":0}))
        for r in rows:
            label = r.get("label")
            if not label: continue
            state = r.get("states")
            CENSUS_DIVS = {
                "New England", "Middle Atlantic",
                "East North Central", "West North Central",
                "South Atlantic", "East South Central", "West South Central",
                "Mountain", "Pacific",
                "NEW ENGLAND", "MID. ATLANTIC", "MIDDLE ATLANTIC",
                "E.N. CENTRAL", "W.N. CENTRAL",
                "S. ATLANTIC", "E.S. CENTRAL", "W.S. CENTRAL",
                "MOUNTAIN", "PACIFIC",
            }
            EXCLUDE_KEYWORDS = ("Total", "Residents", "RESIDENTS", "Non-", "NON-", "Region")
            if (not state) or (state in CENSUS_DIVS) or any(k in state for k in EXCLUDE_KEYWORDS):
                continue
            try:
                year = int(r.get("year",0)); week = int(r.get("week",0))
            except: continue
            cur = safe_int(r.get("m1")); ytd = safe_int(r.get("m3"))
            key = (year, week)
            c = bucket_en[(label, state)][key]
            c["current"] = max(c["current"], cur)  # 同 (label,state,week) は max（重複防止）
            c["cumulative_ytd"] = max(c["cumulative_ytd"], ytd)
            c["year"] = year; c["week"] = week
        for (label, state), wks in bucket_en.items():
            keys = sorted(wks.keys())
            if not keys: continue
            latest = wks[keys[-1]]
            existing = result_en[label].get(state)
            if existing is None or (latest["year"], latest["week"]) > (existing["year"], existing["week"]):
                result_en[label][state] = dict(latest)

        # 修正版: ytd も sum (JP マッピング統合)
        bucket2 = defaultdict(lambda: defaultdict(lambda: {"current":0, "cumulative_ytd":0, "year":None, "week":0}))
        for r in rows:
            label = r.get("label")
            jp = label_to_jp.get(label)
            if not jp: continue
            state = r.get("states")
            # Census divisions / regions / aggregates を除外
            CENSUS_DIVS = {
                "New England", "Middle Atlantic",
                "East North Central", "West North Central",
                "South Atlantic", "East South Central", "West South Central",
                "Mountain", "Pacific",
                # 古い大文字版
                "NEW ENGLAND", "MID. ATLANTIC", "MIDDLE ATLANTIC",
                "E.N. CENTRAL", "W.N. CENTRAL",
                "S. ATLANTIC", "E.S. CENTRAL", "W.S. CENTRAL",
                "MOUNTAIN", "PACIFIC",
            }
            EXCLUDE_KEYWORDS = ("Total", "Residents", "RESIDENTS", "Non-", "NON-", "Region")
            if (not state) or (state in CENSUS_DIVS) or any(k in state for k in EXCLUDE_KEYWORDS):
                continue
            try:
                year = int(r.get("year",0)); week = int(r.get("week",0))
            except: continue
            cur = safe_int(r.get("m1")); ytd = safe_int(r.get("m3"))
            key = (year, week)
            c = bucket2[(jp, state)][key]
            c["current"] += cur
            c["cumulative_ytd"] += ytd  # 同週の異なるラベルの YTD を合算
            c["year"] = year; c["week"] = week
        for (jp, state), wks in bucket2.items():
            keys = sorted(wks.keys())
            if not keys: continue
            latest = wks[keys[-1]]
            # 既存より新しい年/週があれば更新（複数 raw ファイル統合時）
            existing = result[jp].get(state)
            if existing is None or (latest["year"], latest["week"]) > (existing["year"], existing["week"]):
                result[jp][state] = dict(latest)

    print(f"\n集計結果: {len(result)} 疾病 × 平均 {sum(len(v) for v in result.values())/max(len(result),1):.1f} 州")

    # 人口データを CSV から読む
    pop = {}
    if POP_FILE.exists():
        with open(POP_FILE, encoding="utf-8") as f:
            r = csv.DictReader(f)
            for row in r:
                pop[row["state"]] = int(row["population_2024"])
    print(f"州人口データ: {len(pop)} 州")

    # 出力
    PROC_DIR.mkdir(parents=True, exist_ok=True)
    with open(PROC_DIR / "us_state_latest.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  → us_state_latest.json (JPキー): {(PROC_DIR / 'us_state_latest.json').stat().st_size/1024:.0f} KB ({len(result)} 疾病)")
    with open(PROC_DIR / "us_state_latest_en.json", "w", encoding="utf-8") as f:
        json.dump(result_en, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  → us_state_latest_en.json (ENキー): {(PROC_DIR / 'us_state_latest_en.json').stat().st_size/1024:.0f} KB ({len(result_en)} 疾病)")

    with open(PROC_DIR / "us_state_population.json", "w", encoding="utf-8") as f:
        json.dump(pop, f, ensure_ascii=False)
    print(f"  → us_state_population.json: {(PROC_DIR / 'us_state_population.json').stat().st_size/1024:.0f} KB")

    # サンプル表示
    if "麻しん" in result:
        print("\n=== 麻しん 州別 YTD 累積（上位10）===")
        items = sorted(result["麻しん"].items(), key=lambda x: -x[1]["cumulative_ytd"])
        for s, rec in items[:10]:
            p = pop.get(s, 0)
            rate = rec["cumulative_ytd"] / p * 100000 if p else 0
            print(f"  {s:25s} {rec['cumulative_ytd']:>4} 件  ({rate:.2f}/10万)  W{rec['week']}")

    # === Per-disease state × time Excel matrix ===
    # Single source of truth that any downstream tool can consume.
    # Rows = state, columns = year-week, cells = cumulative YTD.
    # Generated only if openpyxl is available (it should be — aggregate_us.py uses it).
    print("\n=== Generating per-disease state × time Excel ===")
    write_state_time_xlsx(label_to_jp, pop)


def write_state_time_xlsx(label_to_jp: dict, population: dict):
    """
    Build per-disease state × week Excel files in us/processed/state_per_disease/.
    Each file has two sheets:
      - "Cumulative YTD" — rows = state, columns = year-week, cells = max(m3) per (state, year, week)
      - "Annual Peak"    — rows = state, columns = year, cells = max(m3) over the year (= year-end count)
    Diseases are JP-merged (multiple NNDSS sub-labels collapsed under one JP name).

    The matrix view makes it trivial for any downstream consumer (dashboard,
    ad-hoc analysis, alerting) to ask "Texas measles by week" or "Utah pertussis
    by year" without re-deriving from raw JSON.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("  [skip] openpyxl not installed — install with: pip install openpyxl --break-system-packages")
        return

    out_dir = PROC_DIR / "state_per_disease"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Re-scan raw state data, this time keeping ALL (year, week) entries
    # (the main aggregator only kept the latest per state).
    # by_disease[jp_disease][state][(year, week)] = cumulative_ytd
    by_disease = defaultdict(lambda: defaultdict(dict))

    CENSUS_DIVS = {
        "New England", "Middle Atlantic", "East North Central", "West North Central",
        "South Atlantic", "East South Central", "West South Central", "Mountain", "Pacific",
        "NEW ENGLAND", "MID. ATLANTIC", "MIDDLE ATLANTIC",
        "E.N. CENTRAL", "W.N. CENTRAL", "S. ATLANTIC", "E.S. CENTRAL", "W.S. CENTRAL",
        "MOUNTAIN", "PACIFIC",
    }
    EXCLUDE_KEYWORDS = ("Total", "Residents", "RESIDENTS", "Non-", "NON-", "Region")

    for path in sorted(RAW_DIR.glob("states_*.json")):
        with open(path, encoding="utf-8") as f:
            rows = json.load(f)
        for r in rows:
            label = r.get("label")
            jp = label_to_jp.get(label)
            if not jp:
                continue
            state = r.get("states")
            if (not state) or (state in CENSUS_DIVS) or any(k in state for k in EXCLUDE_KEYWORDS):
                continue
            try:
                year = int(r.get("year", 0)); week = int(r.get("week", 0))
            except (TypeError, ValueError):
                continue
            ytd = safe_int(r.get("m3"))
            key = (year, week)
            existing = by_disease[jp][state].get(key, 0)
            # Merge multiple NNDSS sub-labels (e.g. "Syphilis, P&S" + "Syphilis, Congenital")
            # by summing their YTD for the same (state, week).
            by_disease[jp][state][key] = existing + ytd

    # Write one workbook per disease
    n_files = 0
    for disease, state_data in sorted(by_disease.items()):
        # Collect all (year, week) keys across this disease's states
        all_keys = set()
        for state, weeks in state_data.items():
            all_keys.update(weeks.keys())
        if not all_keys:
            continue
        sorted_keys = sorted(all_keys)
        sorted_states = sorted(state_data.keys())

        wb = openpyxl.Workbook()

        # Sheet 1: Cumulative YTD by week
        ws1 = wb.active
        ws1.title = "Cumulative YTD"
        ws1.cell(1, 1, "State").font = Font(bold=True)
        for col_idx, (y, w) in enumerate(sorted_keys, start=2):
            c = ws1.cell(1, col_idx, f"{y}-W{w:02d}")
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for row_idx, state in enumerate(sorted_states, start=2):
            ws1.cell(row_idx, 1, state)
            weeks = state_data[state]
            for col_idx, k in enumerate(sorted_keys, start=2):
                ws1.cell(row_idx, col_idx, weeks.get(k, ""))
        ws1.column_dimensions["A"].width = 25
        ws1.freeze_panes = "B2"

        # Sheet 2: Annual peak (= year-end count, the m3 max approach)
        ws2 = wb.create_sheet("Annual Peak")
        years_sorted = sorted({y for (y, _) in sorted_keys})
        ws2.cell(1, 1, "State").font = Font(bold=True)
        for col_idx, y in enumerate(years_sorted, start=2):
            c = ws2.cell(1, col_idx, str(y))
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for row_idx, state in enumerate(sorted_states, start=2):
            ws2.cell(row_idx, 1, state)
            weeks = state_data[state]
            for col_idx, y in enumerate(years_sorted, start=2):
                year_max = 0
                for (yy, ww), v in weeks.items():
                    if yy == y and v > year_max:
                        year_max = v
                ws2.cell(row_idx, col_idx, year_max)
        ws2.column_dimensions["A"].width = 25
        ws2.freeze_panes = "B2"

        # Save with sanitized filename
        out_path = out_dir / f"{_safe_filename(disease)}.xlsx"
        wb.save(out_path)
        n_files += 1

    print(f"  → state_per_disease/: {n_files} ファイル ({len(by_disease)} 疾病 × 平均 {sum(len(v) for v in by_disease.values())/max(len(by_disease),1):.1f} 州)")


if __name__ == "__main__":
    main()
