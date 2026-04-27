"""
Parse JIHS 2024 NESID annual report Excel files (efficient: iter_rows, single pass).
"""
import openpyxl, json, re, os

NESID_DIR = '/sessions/adoring-clever-darwin/mnt/claude/nesid_2024'
OUT = '/sessions/adoring-clever-darwin/nesid_2024.json'

PREFS = {'北海道','青森県','岩手県','宮城県','秋田県','山形県','福島県','茨城県','栃木県','群馬県','埼玉県','千葉県','東京都','神奈川県','新潟県','富山県','石川県','福井県','山梨県','長野県','岐阜県','静岡県','愛知県','三重県','滋賀県','京都府','大阪府','兵庫県','奈良県','和歌山県','鳥取県','島根県','岡山県','広島県','山口県','徳島県','香川県','愛媛県','高知県','福岡県','佐賀県','長崎県','熊本県','大分県','宮崎県','鹿児島県','沖縄県'}

def strip_en(s):
    if not isinstance(s, str): return None
    return re.sub(r'\s*\([^)]*\)\s*', '', s).replace('\u3000', '').strip()

def get_pref_name(raw):
    s = strip_en(raw)
    if not s: return None
    if s.startswith('総') and '数' in s: return '全国'
    return s if s in PREFS else None

def iter_sheet_rows(ws):
    """Yield rows as lists (values). Forward-only."""
    for row in ws.iter_rows(values_only=True):
        yield list(row)

def parse_notifiable_sheet(rows):
    """Return {age_buckets:[...], prefectures:{pref:{age:{total,male,female}}}}."""
    # Find header rows
    r5, r6, data_start = None, None, None
    for i, row in enumerate(rows):
        # Look for 総数(total No.) in this row OR the age header signal
        if r5 is None and any(isinstance(v, str) and '総' in v and '数' in v and 'total' in v.lower() for v in row if v):
            r5 = row
            r6 = rows[i + 1] if i + 1 < len(rows) else None
            data_start = i + 2
            break
    if not r5 or not r6 or data_start is None:
        return None

    # Build col → age mapping
    col_age = {}
    col_sex = {}
    current_age = None
    for c, v in enumerate(r5):
        a = strip_en(v) if v else None
        if a: current_age = a
        sub = strip_en(r6[c]) if c < len(r6) and r6[c] else None
        if current_age and sub in ('総数', '男', '女'):
            col_age[c] = current_age
            col_sex[c] = {'総数':'total','男':'male','女':'female'}[sub]

    # Ordered age buckets (preserve first-seen order)
    age_buckets = []
    seen = set()
    for c in sorted(col_age):
        a = col_age[c]
        if a not in seen:
            age_buckets.append(a); seen.add(a)

    prefectures = {}
    for r in rows[data_start:]:
        if not r: continue
        pref = get_pref_name(r[1] if len(r) > 1 else None)
        if not pref: continue
        pref_data = {a: {'total': None, 'male': None, 'female': None} for a in age_buckets}
        for c, age in col_age.items():
            if c >= len(r): continue
            v = r[c]
            if isinstance(v, (int, float)):
                pref_data[age][col_sex[c]] = v
        prefectures[pref] = pref_data

    return {'age_buckets': age_buckets, 'prefectures': prefectures}

def parse_sentinel_sheet(rows):
    """12-1 has 報告数/定点当たり × 総数/男/女 nested."""
    # Find the METRIC header row: has 3+ occurrences of '定点当たり'
    # (title row only has 1; the actual header row repeats it for every age group)
    r5, r6, r7, data_start = None, None, None, None
    for i, row in enumerate(rows):
        count = sum(1 for v in row if isinstance(v, str) and '定点当たり' in v)
        if count >= 3:
            r5 = rows[i - 1] if i > 0 else None
            r6 = row
            r7 = rows[i + 1] if i + 1 < len(rows) else None
            data_start = i + 2
            break
    if not r5 or not r6 or not r7 or data_start is None:
        return None

    col_age, col_metric, col_sex = {}, {}, {}
    current_age, current_metric = None, None
    for c in range(max(len(r5), len(r6), len(r7))):
        v5 = r5[c] if c < len(r5) else None
        v6 = r6[c] if c < len(r6) else None
        v7 = r7[c] if c < len(r7) else None
        a = strip_en(v5) if v5 else None
        if a: current_age = a
        m_raw = strip_en(v6) if v6 else None
        if m_raw and '報告数' in m_raw: current_metric = 'cases'
        elif m_raw and ('定点当たり' in m_raw or '定点当り' in m_raw): current_metric = 'per_sentinel'
        sub = strip_en(v7) if v7 else None
        if current_age and current_metric and sub in ('総数','男','女'):
            col_age[c] = current_age
            col_metric[c] = current_metric
            col_sex[c] = {'総数':'total','男':'male','女':'female'}[sub]

    age_buckets = []
    seen = set()
    for c in sorted(col_age):
        a = col_age[c]
        if a not in seen:
            age_buckets.append(a); seen.add(a)

    prefectures = {}
    for r in rows[data_start:]:
        if not r: continue
        pref = get_pref_name(r[1] if len(r) > 1 else None)
        if not pref: continue
        pref_data = {a: {'cases': {'total': None,'male':None,'female':None},
                         'per_sentinel': {'total':None,'male':None,'female':None}} for a in age_buckets}
        for c, age in col_age.items():
            if c >= len(r): continue
            v = r[c]
            if isinstance(v, (int, float)):
                metric = col_metric[c]
                if metric == 'per_sentinel':
                    v = round(float(v), 3)
                pref_data[age][metric][col_sex[c]] = v
        prefectures[pref] = pref_data
    return {'age_buckets': age_buckets, 'prefectures': prefectures}


def parse_file(path, parser):
    print(f'Parsing {os.path.basename(path)}...', flush=True)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    for i, sheet_name in enumerate(wb.sheetnames):
        if i % 20 == 0:
            print(f'  [{i+1}/{len(wb.sheetnames)}] {sheet_name}', flush=True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        result = parser(rows)
        if result:
            out[sheet_name] = result
    wb.close()
    print(f'  Done: {len(out)} diseases', flush=True)
    return out


notifiable = parse_file(f'{NESID_DIR}/Syu_4_1.xlsx', parse_notifiable_sheet)
notifiable_det = parse_file(f'{NESID_DIR}/Syu_4_2.xlsx', parse_notifiable_sheet)
sentinel = parse_file(f'{NESID_DIR}/Syu_12_1.xlsx', parse_sentinel_sheet)

out = {
    'metadata': {
        'year': 2024,
        'source': 'https://id-info.jihs.go.jp/surveillance/idwr/annual/2024/',
        'unit': 'persons (cases). Sentinel also has per-sentinel values.',
        'parsed': '2026-04-17',
    },
    'notifiable': notifiable,
    'notifiable_detailed': notifiable_det,
    'sentinel': sentinel,
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
print(f'\nSaved: {OUT} ({os.path.getsize(OUT):,} bytes)')

# Sanity
k = notifiable.get('結核', {}).get('prefectures', {}).get('全国', {}).get('総数', {})
i = sentinel.get('インフルエンザ', {}).get('prefectures', {}).get('全国', {}).get('総数', {})
print(f'結核 全国 総数: {k}')
print(f'インフル 全国 総数: {i}')
