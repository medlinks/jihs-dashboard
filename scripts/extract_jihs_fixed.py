"""
IDWR 感染症週報 PDFからデータ抽出 → Excel出力
Usage:
  python extract_idwr.py --pdf-dir /path/to/pdfs --year 2014 --output /path/to/out.xlsx

行: 都道府県 (総数 + 47都道府県)
列: 週 (第1-2週 〜 第52週)
シート: 1疾病1シート（報告数のみ）

Dependencies:
  pip install pdfplumber openpyxl --break-system-packages
"""

import argparse, pdfplumber, re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Constants ────────────────────────────────────────────────────────────────

PREFECTURES = [
    "総数","北海道","青森県","岩手県","宮城県","秋田県","山形県","福島県",
    "茨城県","栃木県","群馬県","埼玉県","千葉県","東京都","神奈川県",
    "新潟県","富山県","石川県","福井県","山梨県","長野県",
    "岐阜県","静岡県","愛知県","三重県",
    "滋賀県","京都府","大阪府","兵庫県","奈良県","和歌山県",
    "鳥取県","島根県","岡山県","広島県","山口県",
    "徳島県","香川県","愛媛県","高知県",
    "福岡県","佐賀県","長崎県","熊本県","大分県","宮崎県","鹿児島県","沖縄県"
]

NUM_PAT  = re.compile(r'^[\d,]+\.?\d*$|^-$')
SKIP_ROW = {'Infectious', 'Ministry', '感染症週報', '発行', '厚生', '国立', '注）'}
SKIP_WORD = {'*鳥', '注）', 'は除く', 'Infectious', 'Ministry', '掲載', 'インフルエンザは', '新型インフルエンザ'}

# Excel styling
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")
THIN        = Side(style='thin')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Core extraction helpers ──────────────────────────────────────────────────

def group_rows(words, y_tol=4):
    """Group words into horizontal rows by vertical position."""
    rows = []
    for w in sorted(words, key=lambda w: w['top']):
        placed = False
        for row in rows:
            if abs(w['top'] - row[0]) <= y_tol:
                row[1].append(w)
                placed = True
                break
        if not placed:
            rows.append([w['top'], [w]])
    return rows


def voronoi_assign(word_x0, report_xs, left_bias=3):
    """
    Assign a word to the nearest column center using Voronoi boundaries.
    left_bias: subtract this many pixels before comparison. Fixes characters
    that fall just outside their correct Voronoi cell due to PDF rendering
    (e.g. '痘' in '水痘' falling 0.28px right of boundary).
    """
    adj = word_x0 - left_bias
    xs = sorted(report_xs)
    for i, cx in enumerate(xs):
        lft = (xs[i-1] + cx) / 2 if i > 0 else float('-inf')
        rgt = (xs[i+1] + cx) / 2 if i < len(xs)-1 else float('inf')
        if lft <= adj < rgt:
            return cx
    return min(xs, key=lambda cx: abs(cx - word_x0))


def find_pref(wds):
    """Find the prefecture name in a row of words."""
    full = ''.join(w['text'] for w in wds)
    for p in PREFECTURES:
        if p in full:
            return p
    return None


def extract_disease_names(rows, col_row_idx, report_xs):
    """
    Extract disease names from the rows above the column header row.
    Each disease name may span multiple rows and characters may be split.
    Returns a list of names aligned with report_xs (sorted).
    """
    col_to_words = defaultdict(list)
    for i in range(max(0, col_row_idx - 4), col_row_idx):
        _, wds = rows[i]
        row_txt = ''.join(w['text'] for w in wds)
        # Skip rows that are clearly page headers, not disease names
        if any(s in row_txt for s in [
            'Infectious', 'Ministry', '都道府県', '疾病', '報告数',
            '通巻', '注）', 'は除く', '掲載', '新型インフルエンザ'
        ]):
            continue
        for w in wds:
            if any(s in w['text'] for s in SKIP_WORD):
                continue
            cx = voronoi_assign(w['x0'], report_xs)
            col_to_words[cx].append(w)

    names = []
    for cx in sorted(report_xs):
        wds = sorted(col_to_words.get(cx, []),
                     key=lambda w: (round(w['top'] / 8) * 8, w['x0']))
        name = ''.join(w['text'] for w in wds).replace('*', '').strip()
        # Clean footnote text: keep only what comes after the last sentence period
        if '。' in name:
            name = name.split('。')[-1].strip()
        # Strip "YYYY年第X週" / "YYYY年第X-Y週" prefix (2023+ PDFs print week label
        # directly above the disease column, causing it to merge into the name)
        import re as _re2
        name = _re2.sub(r'^\d{4}年第\d+(?:-\d+)?週', '', name).strip()
        names.append(name if name else f"疾病@{int(cx)}")
    return names


def extract_page(page):
    """
    Extract disease→prefecture→count data from a single PDF page.
    Returns: {disease_name: {pref: count_or_None}}
    """
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    if not words:
        return {}
    rows = group_rows(words)
    page_text = ''.join(w['text'] for w in words)

    # ── Special case: インフルエンザ入院患者 (single-disease hospitalization page)
    if '入院患者' in page_text and '報告数・疾病・都道府県別' in page_text:
        result = defaultdict(dict)
        dname = 'インフルエンザ(入院患者)'
        in_data = False
        for _, wds in rows:
            txt = ''.join(w['text'] for w in wds)
            if any(s in txt for s in SKIP_ROW):
                continue
            texts = [w['text'] for w in wds]
            if '報告数' in texts and '定点' not in txt and '累積' not in txt:
                in_data = True
                continue
            if not in_data:
                continue
            pref = find_pref(wds)
            if pref:
                vals = [w['text'].replace(',', '') for w in wds
                        if NUM_PAT.match(w['text'])]
                if vals:
                    v = vals[0]
                    result[dname][pref] = None if v == '-' else int(v)
        return dict(result)

    # ── Standard multi-column page: find header row
    col_row_idx, col_words = None, None
    for i, (_, wds) in enumerate(rows):
        texts = [w['text'] for w in wds]
        if '報告数' in texts and ('累積' in texts or '定点当り' in texts):
            col_row_idx, col_words = i, wds
            break
    if col_row_idx is None:
        return {}

    # Parse column positions and types (報告数 / 累積 / 定点当り)
    col_types = [
        (w['x0'], w['text'])
        for w in sorted(col_words, key=lambda w: w['x0'])
        if w['text'] in ('報告数', '累積', '定点当り')
    ]
    report_xs = [x for x, ct in col_types if ct == '報告数']
    # ri_list: ordered index of each 報告数 column among all column types
    # This lets us pick the correct value from each data row by position
    ri_list = [i for i, (x, ct) in enumerate(col_types) if ct == '報告数']

    if not report_xs:
        return {}

    disease_names = extract_disease_names(rows, col_row_idx, report_xs)

    result = defaultdict(dict)
    for row_idx, (_, wds) in enumerate(rows):
        if row_idx <= col_row_idx:
            continue
        txt = ''.join(w['text'] for w in wds)
        if any(s in txt for s in SKIP_ROW):
            continue
        pref = find_pref(wds)
        if not pref:
            continue
        vals = [
            w['text'].replace(',', '')
            for w in sorted(wds, key=lambda w: w['x0'])
            if NUM_PAT.match(w['text'])
        ]
        for d_idx, ri in enumerate(ri_list):
            if ri < len(vals) and d_idx < len(disease_names):
                v = vals[ri]
                val = None if v == '-' else (int(v) if '.' not in v else float(v))
                result[disease_names[d_idx]][pref] = val
    return dict(result)


# ── File utilities ───────────────────────────────────────────────────────────

def week_label(pdf_path, year):
    """
    Extract week label from filename.
    idwr2014-03.pdf     → 第3週
    idwr2014-17-18.pdf  → 第17-18週
    """
    m = re.search(rf'idwr{year}-(\d+)(?:-(\d+))?\.pdf', pdf_path.name, re.IGNORECASE)
    if not m:
        # Fallback: grab first number sequence
        m2 = re.search(r'(\d+)(?:-(\d+))?', pdf_path.stem)
        if not m2:
            return pdf_path.stem
        w1, w2 = int(m2.group(1)), m2.group(2)
    else:
        w1, w2 = int(m.group(1)), m.group(2)
    return f"第{w1}-{int(w2)}週" if w2 else f"第{w1}週"


def clean_disease_name(name):
    """Ensure sheet name is clean and within Excel's 31-char limit."""
    name = name.strip()
    if '。' in name:
        name = name.split('。')[-1].strip()
    return name[:31]


def should_skip_page(page_text):
    """Return True for pages that should not be processed."""
    # Gender-breakdown STI monthly pages
    if '性別（男）' in page_text or '性別（女）' in page_text:
        return True
    if '性器クラミジア' in page_text:
        return True
    # Veterinary pages (pure animal surveillance)
    if '獣医師' in page_text:
        return True
    # Combined human/animal surveillance page for MERS/bird-flu (2015+)
    # This page shows 鳥インフルエンザ + 中東呼吸器症候群 broken down by host
    # species (サル, 鳥類, ヒトコブラクダ etc.) — NOT individual human case data.
    # Human H5N1/H7N9 cases are reported separately on the 全数報告 page.
    if 'ヒトコブラクダ' in page_text:
        return True
    return False


# ── Main extraction ──────────────────────────────────────────────────────────

def extract_year(pdf_dir: Path, year: int):
    """
    Process all IDWR PDFs for a given year directory.
    Returns (all_data, week_labels):
      all_data: {disease: {pref: {week: value}}}
      week_labels: ordered list of week strings
    """
    pattern = f"idwr{year}-*.pdf"
    pdf_files = sorted(
        pdf_dir.glob(pattern),
        key=lambda p: int(re.search(r'(\d+)', p.stem).group(1))
    )
    if not pdf_files:
        raise FileNotFoundError(
            f"No files matching '{pattern}' found in {pdf_dir}\n"
            f"Files present: {list(pdf_dir.iterdir())[:10]}"
        )

    all_data = defaultdict(lambda: defaultdict(dict))
    week_labels = []

    for pdf_path in pdf_files:
        # Detect combined PDFs (e.g. idwr2024-17-18.pdf → w1=17, w2=18)
        m = re.search(rf'idwr{year}-(\d+)(?:-(\d+))?\.pdf', pdf_path.name, re.IGNORECASE)
        if not m:
            m = re.search(r'(\d+)(?:-(\d+))?', pdf_path.stem)
        w1_str = m.group(1) if m else None
        w2_str = m.group(2) if m else None
        is_combined = bool(w2_str)

        if is_combined:
            w1_num, w2_num = int(w1_str), int(w2_str)
            wk1, wk2 = f"第{w1_num}週", f"第{w2_num}週"
            week_labels.extend([wk1, wk2])
            print(f"  {pdf_path.name} → {wk1} + {wk2}", flush=True)

            # Pages before the first week marker are skipped (e.g. prev-month data)
            current_wk = None
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    words = page.extract_words(x_tolerance=3, y_tolerance=3)
                    page_text = ''.join(w['text'] for w in words)

                    # Detect section transitions: "1週のデータ" / "第17週のデータ" etc.
                    # Check w2 first so a page with both markers (shouldn't happen) stays in w2
                    if f'{w2_num}週のデータ' in page_text:
                        current_wk = wk2
                    elif f'{w1_num}週のデータ' in page_text:
                        current_wk = wk1

                    if current_wk is None:
                        continue  # skip intro / prev-month pages
                    if '報告数' not in page_text or '北海道' not in page_text:
                        continue
                    if should_skip_page(page_text):
                        continue

                    page_data = extract_page(page)
                    for disease, pref_data in page_data.items():
                        dname = clean_disease_name(disease)
                        for pref, val in pref_data.items():
                            # Store 0 for '-' entries so all-zero diseases still
                            # get an Excel sheet (important for rare Class 1 diseases
                            # like ジフテリア / 鳥インフルエンザH5N1 / MERS that need
                            # to be tracked even when no cases are reported)
                            all_data[dname][pref][current_wk] = 0 if val is None else val
        else:
            wk = week_label(pdf_path, year)
            week_labels.append(wk)
            print(f"  {pdf_path.name} → {wk}", flush=True)

            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    words = page.extract_words(x_tolerance=3, y_tolerance=3)
                    page_text = ''.join(w['text'] for w in words)

                    if '報告数' not in page_text or '北海道' not in page_text:
                        continue
                    if should_skip_page(page_text):
                        continue

                    page_data = extract_page(page)
                    for disease, pref_data in page_data.items():
                        dname = clean_disease_name(disease)
                        for pref, val in pref_data.items():
                            all_data[dname][pref][wk] = 0 if val is None else val

    print(f"  → {len(all_data)} diseases found across {len(week_labels)} weeks")
    return dict(all_data), week_labels


# ── Excel builder ────────────────────────────────────────────────────────────

def build_excel(all_data, week_labels, output_path):
    """Write all_data to a formatted Excel file."""
    wb = Workbook()
    wb.remove(wb.active)

    import re as _re
    def _sanitize(s):
        """Strip control characters and zero-width characters (PDF artifacts)."""
        # Remove ASCII control chars + zero-width Unicode chars
        return _re.sub(r'[\x00-\x1f\x7f\u200b\u200c\u200d\ufeff]', '', s)

    used_names = {}
    for disease in sorted(all_data.keys()):
        raw = _sanitize(disease)[:31]
        if raw in used_names:
            used_names[raw] += 1
            title = raw[:28] + str(used_names[raw])
        else:
            used_names[raw] = 0
            title = raw

        ws = wb.create_sheet(title=title)

        # Header row
        ws.cell(1, 1, "都道府県").font = HEADER_FONT
        ws.cell(1, 1).fill = HEADER_FILL
        ws.cell(1, 1).alignment = Alignment(horizontal='center')
        for j, wk in enumerate(week_labels, start=2):
            c = ws.cell(1, j, wk)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal='center')

        # Data rows
        for i, pref in enumerate(PREFECTURES, start=2):
            fill = ALT_FILL if i % 2 == 0 else None
            ws.cell(i, 1, pref)
            if fill:
                ws.cell(i, 1).fill = fill
            for j, wk in enumerate(week_labels, start=2):
                val = all_data[disease].get(pref, {}).get(wk)
                cell = ws.cell(i, j, val)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal='right')

        # Borders and freeze panes
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                  min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = BORDER
        ws.freeze_panes = "B2"

        # Column widths
        ws.column_dimensions['A'].width = 14
        for j in range(2, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, j).column_letter].width = 9

    wb.save(output_path)
    print(f"\nSaved: {output_path}  ({len(wb.sheetnames)} sheets)")
    return wb


# ── Verification helper ──────────────────────────────────────────────────────

KNOWN_VALUES = {
    2013: {('インフルエンザ', '総数', '第3週'): 111475,
           ('結核',          '総数', '第3週'): 276,
           ('水痘',          '総数', '第3週'): 3539},
    2014: {('インフルエンザ', '総数', '第3週'): 58233,
           ('結核',          '総数', '第3週'): 263,
           ('水痘',          '総数', '第3週'): 3462},
}

def verify(wb, year):
    """Check known reference values for the given year."""
    refs = KNOWN_VALUES.get(year, {})
    if not refs:
        print("(No reference values for this year — skipping verification)")
        return
    all_ok = True
    for (disease, pref, week), expected in refs.items():
        if disease not in wb.sheetnames:
            print(f"  ✗ '{disease}': sheet not found"); all_ok = False; continue
        ws = wb[disease]
        week_col = next((c for c in range(2, ws.max_column+1)
                         if ws.cell(1,c).value == week), None)
        pref_row = next((r for r in range(2, ws.max_row+1)
                         if ws.cell(r,1).value == pref), None)
        val = ws.cell(pref_row, week_col).value if week_col and pref_row else None
        ok = val == expected
        mark = "✓" if ok else "✗"
        print(f"  {mark} {disease} / {pref} / {week}: {val} (expected {expected})")
        if not ok:
            all_ok = False
    if all_ok:
        print("  All reference values match!")


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extract IDWR 感染症週報 PDF data to Excel"
    )
    parser.add_argument('--pdf-dir',  required=True, help="Directory containing IDWR PDFs")
    parser.add_argument('--year',     required=True, type=int, help="Year (e.g. 2014)")
    parser.add_argument('--output',   required=True, help="Output Excel file path (.xlsx)")
    parser.add_argument('--no-verify', action='store_true', help="Skip verification step")
    args = parser.parse_args()

    pdf_dir = Path(args.pdf_dir)
    if not pdf_dir.exists():
        raise SystemExit(f"PDF directory not found: {pdf_dir}")

    print(f"Extracting IDWR {args.year} from {pdf_dir} ...")
    all_data, week_labels = extract_year(pdf_dir, args.year)
    wb = build_excel(all_data, week_labels, args.output)

    if not args.no_verify:
        print("\nVerification:")
        verify(wb, args.year)


if __name__ == '__main__':
    main()
