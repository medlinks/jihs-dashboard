"""A2: Extract 12-disease prefecture × week timeseries from IDWR{year}.xlsx (2013-2026).

Schema: disease, year, iso_week, prefecture, cases
Missing data → omitted (not zero-filled).

Output:
  prefecture_week_timeseries.csv  — long format, ~12 × 47 × 700 = 394,800 rows
  prefecture_week_timeseries_summary.md — coverage / sanity check report
"""
import openpyxl
import csv
import json
import re
from pathlib import Path

WE_DIR = Path('/sessions/cool-clever-goldberg/mnt/claude/weekly_extracted')
OUT_DIR = Path('/sessions/cool-clever-goldberg/mnt/outputs')

TARGETS = {
    # IDWR sheet name → label
    '手足口病':                    'sentinel',
    'RSウイルス感染症':            'sentinel',
    'インフルエンザ':              'sentinel',
    'マイコプラズマ肺炎':          'sentinel',
    'ヘルパンギーナ':              'sentinel',
    '感染性胃腸炎':                'sentinel',
    '流行性耳下腺炎':              'sentinel',
    'Ａ群溶血性レンサ球菌咽頭炎':  'sentinel',  # full-width Ａ
    '咽頭結膜熱':                  'sentinel',
    '風しん':                      'full-report',
    '麻しん':                      'full-report',
    '梅毒':                        'full-report',
}

VALID_PREFS = {
    '北海道','青森県','岩手県','宮城県','秋田県','山形県','福島県',
    '茨城県','栃木県','群馬県','埼玉県','千葉県','東京都','神奈川県',
    '新潟県','富山県','石川県','福井県','山梨県','長野県',
    '岐阜県','静岡県','愛知県','三重県',
    '滋賀県','京都府','大阪府','兵庫県','奈良県','和歌山県',
    '鳥取県','島根県','岡山県','広島県','山口県',
    '徳島県','香川県','愛媛県','高知県',
    '福岡県','佐賀県','長崎県','熊本県','大分県','宮崎県','鹿児島県','沖縄県'
}

def main():
    files = sorted(WE_DIR.glob('IDWR*.xlsx'))
    print(f'Found {len(files)} IDWR XLSX files: {[f.name for f in files]}\n')

    all_rows = []
    coverage = {d: {} for d in TARGETS}  # disease → year → set of (week, n_prefs_with_data)
    missing_disease_year = []

    for fp in files:
        year = int(re.search(r'IDWR(\d{4})', fp.name).group(1))
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        sheet_set = set(wb.sheetnames)
        for disease in TARGETS:
            if disease not in sheet_set:
                missing_disease_year.append((disease, year, fp.name))
                continue
            ws = wb[disease]
            rows = list(ws.iter_rows(values_only=True))
            if not rows: continue
            header = rows[0]  # ['都道府県', '第1週', '第2週', ...]
            week_nums = []
            for col in header[1:]:
                if col is None:
                    week_nums.append(None); continue
                m = re.search(r'第(\d+)', str(col))
                week_nums.append(int(m.group(1)) if m else None)
            year_weeks_seen = set()
            year_pref_counts = {}  # week → n_prefs_with_data
            for row in rows[1:]:
                pref = row[0]
                if pref is None: continue
                pref = str(pref).strip()
                # Skip 総数 (national total) and non-prefecture rows; keep only valid 47
                if pref not in VALID_PREFS:
                    continue
                for col_idx, val in enumerate(row[1:]):
                    if col_idx >= len(week_nums): break
                    w = week_nums[col_idx]
                    if w is None or val is None: continue
                    try:
                        cases = int(val)
                    except (TypeError, ValueError):
                        try: cases = int(float(val))
                        except: continue
                    all_rows.append({
                        'disease': disease,
                        'year': year, 'iso_week': w,
                        'prefecture': pref, 'cases': cases,
                    })
                    year_weeks_seen.add(w)
                    year_pref_counts[w] = year_pref_counts.get(w, 0) + 1
            # Coverage stats
            coverage[disease][year] = {
                'weeks_with_data': len(year_weeks_seen),
                'avg_prefs_per_week': sum(year_pref_counts.values()) / max(1, len(year_pref_counts)),
                'min_week': min(year_weeks_seen) if year_weeks_seen else None,
                'max_week': max(year_weeks_seen) if year_weeks_seen else None,
            }
        wb.close()
        print(f'  {fp.name}: processed {sum(1 for r in all_rows if r["year"]==year):,} rows')

    print(f'\nTotal rows extracted: {len(all_rows):,}')

    # Save CSV
    csv_out = OUT_DIR / 'prefecture_week_timeseries.csv'
    with open(csv_out, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['disease','year','iso_week','prefecture','cases'])
        writer.writeheader()
        writer.writerows(all_rows)
    print(f'Saved: {csv_out} ({csv_out.stat().st_size:,} bytes)')

    # Save coverage JSON
    cov_out = OUT_DIR / 'prefecture_week_timeseries_coverage.json'
    with open(cov_out, 'w', encoding='utf-8') as f:
        json.dump({
            'coverage': coverage,
            'missing_disease_year': missing_disease_year,
            'total_rows': len(all_rows),
        }, f, ensure_ascii=False, indent=2)
    print(f'Saved: {cov_out}')

    # Print coverage summary
    print(f'\n=== Coverage summary (rows per disease/year) ===')
    print(f'{"Disease":<28} ', end='')
    years = sorted(set(y for d in coverage.values() for y in d))
    for y in years: print(f'{y:>5}', end='')
    print()
    for d in TARGETS:
        print(f'{d:<28} ', end='')
        for y in years:
            ww = coverage[d].get(y, {}).get('weeks_with_data', 0)
            print(f'{ww:>5}', end='')
        print()

    # Sanity: print latest week values for each disease
    print(f'\n=== Sanity: most recent week values (2026 W16) ===')
    for d in TARGETS:
        latest_rows = [r for r in all_rows if r['disease']==d and r['year']==2026 and r['iso_week']==16]
        n_prefs = len(latest_rows)
        total = sum(r['cases'] for r in latest_rows)
        # Compare to dashboard.html DATA.weekly_trends (national total)
        with open('/sessions/cool-clever-goldberg/mnt/claude/scripts/full_dashboard_data.json') as f:
            dash = json.load(f)
        dash_total = None
        if d in dash['weekly_trends']:
            for r in dash['weekly_trends'][d]:
                if r['year'] == 2026 and r['iso_week' if 'iso_week' in r else 'week'] == 16:
                    dash_total = r['total']; break
            if dash_total is None:
                for r in dash['weekly_trends'][d]:
                    if r['year'] == 2026 and r.get('week') == 16:
                        dash_total = r['total']; break
        print(f'  {d:<28} 47-pref sum={total:>7,} (n_pref={n_prefs:2d}) | dashboard national={dash_total}')

    if missing_disease_year:
        print(f'\n=== Missing disease/year combos: {len(missing_disease_year)} ===')
        for d, y, fn in missing_disease_year[:20]:
            print(f'  {d} not in {fn}')

if __name__ == '__main__':
    main()
