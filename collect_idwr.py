"""
IDWR 自動収集スクリプト
毎週火曜日にCoworkスケジューラーから実行する

処理フロー:
1. updates.html をチェックして未処理の速報があるか確認
2. 新しい速報があれば zensu.csv をダウンロード
3. 87疾病のExcelファイルに新週のデータを追記
4. 実行ログを保存
"""

import json
import re
import ssl
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# パス設定
# ============================================================
BASE_DIR     = Path(__file__).parent
DISEASES_DIR = BASE_DIR / 'diseases'
STATE_FILE   = BASE_DIR / 'state.json'
LOG_FILE     = BASE_DIR / 'collect_log.txt'

UPDATES_URL  = 'https://id-info.jihs.go.jp/updates.html'
IDWR_BASE    = 'https://id-info.jihs.go.jp/surveillance/idwr/provisional'

# ============================================================
# SSL設定（証明書エラー回避）
# ============================================================
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

def fetch(url: str) -> str:
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, context=SSL_CTX) as r:
        return r.read().decode('utf-8', errors='replace')

def fetch_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, context=SSL_CTX) as r:
        return r.read()

def log(msg: str):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f'[{ts}] {msg}'
    print(line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')

# ============================================================
# 状態管理
# ============================================================
def load_state() -> dict:
    if STATE_FILE.exists():
        with open(STATE_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {'processed_weeks': []}

def save_state(state: dict):
    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

# ============================================================
# updates.html から未処理の速報を検出
# ============================================================
def find_new_reports(state: dict) -> list:
    """
    updates.html を解析して未処理の IDWR速報データ を返す
    戻り値: [{'year': 2026, 'week': 15, 'label': '...', 'url': '...'}, ...]
    """
    log('updates.html をチェック中...')
    html = fetch(UPDATES_URL)

    # aタグから IDWR速報データ のリンクを抽出
    pattern = re.compile(
        r'<a[^>]+href="([^"]*provisional/(\d{4})/(\d+)/index\.html)"[^>]*>'
        r'[^<]*(\d{4})年(\d+)週[^<]*速報[^<]*</a>',
        re.DOTALL
    )

    # より柔軟なパターン
    link_pattern = re.compile(r'href="(\.?/?[^"]*provisional/(\d{4})/(\d+)/index\.html)"')
    title_pattern = re.compile(r'IDWR速報データ\s*(\d{4})年第?(\d+)週')

    found = []

    # リンクとタイトルを同時に抽出
    links_raw = re.findall(
        r'<a[^>]+href="([^"]*provisional/(\d{4})/(\d+)/index\.html)"[^>]*>(.*?)</a>',
        html, re.DOTALL
    )

    for href, year_str, week_str, text in links_raw:
        year = int(year_str)
        week = int(week_str)
        week_key = f'{year}-W{week:02d}'

        if week_key not in state['processed_weeks']:
            # 絶対URLに変換
            if href.startswith('./'):
                url = 'https://id-info.jihs.go.jp' + href[1:]
            elif href.startswith('/'):
                url = 'https://id-info.jihs.go.jp' + href
            else:
                url = href

            clean_text = re.sub(r'<[^>]+>', '', text).strip()
            found.append({
                'year': year,
                'week': week,
                'key': week_key,
                'label': clean_text,
                'url': url
            })
            log(f'  未処理の速報を発見: {year}年第{week}週')

    if not found:
        log('  新しい速報はありません')

    return found

# ============================================================
# 速報ページから zensu.csv の URL を取得してダウンロード
# ============================================================
def download_zensu(report: dict) -> tuple[bytes, str] | None:
    """
    週報ページを開いて zensu.csv をダウンロード
    戻り値: (csvバイト列, csvURL) または None
    """
    year  = report['year']
    week  = report['week']
    page_url = f'{IDWR_BASE}/{year}/{week}/index.html'

    log(f'  速報ページを取得: {page_url}')
    try:
        html = fetch(page_url)
    except Exception as e:
        log(f'  ERROR ページ取得失敗: {e}')
        return None

    # zensu.csv のリンクを探す
    # パターン: href="./YYYY-WW-zensu.csv" または類似
    csv_match = re.search(r'href="(\./)?([^"]*zensu[^"]*\.csv)"', html)
    if not csv_match:
        log(f'  ERROR zensu.csv リンクが見つかりません')
        return None

    csv_filename = csv_match.group(2).lstrip('./')
    csv_url = f'{IDWR_BASE}/{year}/{week}/{csv_filename}'
    log(f'  zensu.csv ダウンロード: {csv_url}')

    try:
        csv_bytes = fetch_bytes(csv_url)
        log(f'  ダウンロード完了: {len(csv_bytes):,} bytes')
        return csv_bytes, csv_url
    except Exception as e:
        log(f'  ERROR ダウンロード失敗: {e}')
        return None

# ============================================================
# zensu.csv のパース
# ============================================================
def parse_zensu(csv_bytes: bytes) -> dict | None:
    """
    zensu.csv を解析して構造化データを返す
    戻り値:
    {
        'year': 2026, 'week': 14, 'period': '03月30日〜04月05日',
        'created': '2026年04月08日作成',
        'diseases': [{'name': 'エボラ出血熱', 'col_idx': 1}, ...],
        'prefectures': ['総数', '北海道', ...],
        'data': {'都道府県名': ['col0', 'col1', ...], ...}
    }
    """
    try:
        text = csv_bytes.decode('shift_jis')
    except Exception as e:
        log(f'  ERROR CSV デコード失敗: {e}')
        return None

    lines = text.splitlines()
    if len(lines) < 5:
        log(f'  ERROR CSV 行数不足: {len(lines)}行')
        return None

    # 行2: メタ情報
    meta = lines[1].split(',')
    week_label   = meta[0].strip()
    created_date = meta[1].strip() if len(meta) > 1 else ''

    week_m  = re.search(r'(\d{4})年(\d+)週', week_label)
    period_m = re.search(r'\((.+?)\)', week_label)

    if not week_m:
        log(f'  ERROR 週番号が解析できません: {week_label}')
        return None

    year     = int(week_m.group(1))
    week_num = int(week_m.group(2))
    period   = period_m.group(1) if period_m else ''

    # 行3: 疾病名
    row3 = lines[2].split(',')
    diseases = [
        {'name': v.strip(), 'col_idx': i}
        for i, v in enumerate(row3) if v.strip()
    ]

    # 行5以降: 都道府県データ
    prefectures = []
    data = {}
    for line in lines[4:]:
        cols = line.split(',')
        pref = cols[0].strip()
        if pref:
            prefectures.append(pref)
            data[pref] = cols

    return {
        'year': year, 'week': week_num,
        'period': period, 'created': created_date,
        'diseases': diseases,
        'prefectures': prefectures,
        'data': data
    }

# ============================================================
# 87疾病の Excel に新週データを追記
# ============================================================
def safe_dirname(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', name)

def get_val(row_data: list, col_idx: int) -> int:
    try:
        return int(row_data[col_idx].strip())
    except (IndexError, ValueError):
        return 0

# スタイル定義
def _make_styles():
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font  = Font(name='Arial', size=9)
    zero_font  = Font(name='Arial', size=9, color='AAAAAA')
    total_font = Font(name='Arial', bold=True, size=9)
    total_fill = PatternFill('solid', start_color='FFF2CC')
    right = Alignment(horizontal='right', vertical='center')
    center = Alignment(horizontal='center', vertical='center')
    return border, data_font, zero_font, total_font, total_fill, right, center

STYLES = _make_styles()

def append_week_to_excel(xlsx_path: Path, disease: dict, parsed: dict):
    """
    既存の Excel ファイルを開き、新週の行を追記する
    """
    if not xlsx_path.exists():
        log(f'    WARNING Excelファイルが存在しません: {xlsx_path.name}')
        return False

    col_report = disease['col_idx']
    col_cumul  = disease['col_idx'] + 1
    FIXED_COLS = 4

    border, data_font, zero_font, total_font, total_fill, right, center = STYLES

    try:
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb['速報データ']

        # 次の空行を探す（5行目から）
        next_row = ws.max_row + 1
        if next_row < 5:
            next_row = 5

        # 重複チェック：同じ年・週がすでに存在しないか
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0] == parsed['year'] and row[1] == parsed['week']:
                log(f'    SKIP {xlsx_path.name}: {parsed["year"]}年第{parsed["week"]}週は既に記録済み')
                wb.close()
                return False

        # 固定列（年・週・期間・作成日）
        for ci, val in enumerate([
            parsed['year'], parsed['week'],
            parsed['period'], parsed['created']
        ], start=1):
            cell = ws.cell(row=next_row, column=ci, value=val)
            cell.font = data_font
            cell.alignment = center
            cell.border = border

        # 都道府県データ列
        for pi, pref in enumerate(parsed['prefectures']):
            row_data = parsed['data'].get(pref, [])
            val_rep  = get_val(row_data, col_report)
            val_cum  = get_val(row_data, col_cumul)

            c_rep = FIXED_COLS + pi * 2 + 1
            c_cum = FIXED_COLS + pi * 2 + 2

            for c, val in [(c_rep, val_rep), (c_cum, val_cum)]:
                cell = ws.cell(row=next_row, column=c, value=val)
                cell.font = zero_font if val == 0 else data_font
                cell.alignment = right
                cell.border = border

            # 総数行は背景色
            if pref == '総数':
                for c in [c_rep, c_cum]:
                    ws.cell(row=next_row, column=c).fill = total_fill
                    ws.cell(row=next_row, column=c).font = total_font

        # 疾病情報シートの最終更新日を更新
        if '疾病情報' in wb.sheetnames:
            wb['疾病情報']['B5'] = parsed['created']

        wb.save(str(xlsx_path))
        wb.close()
        return True

    except Exception as e:
        log(f'    ERROR Excel更新失敗 {xlsx_path.name}: {e}')
        return False

# ============================================================
# メイン処理
# ============================================================
def main():
    log('=' * 60)
    log('IDWR 自動収集スクリプト 開始')
    log('=' * 60)

    state = load_state()
    new_reports = find_new_reports(state)

    if not new_reports:
        log('処理完了: 新しい速報なし')
        return

    results = []

    for report in sorted(new_reports, key=lambda x: (x['year'], x['week'])):
        year = report['year']
        week = report['week']
        log(f'\n--- {year}年第{week}週 処理開始 ---')

        # zensu.csv ダウンロード
        result = download_zensu(report)
        if result is None:
            log(f'  SKIP {year}年第{week}週: ダウンロード失敗')
            continue
        csv_bytes, csv_url = result

        # CSV パース
        parsed = parse_zensu(csv_bytes)
        if parsed is None:
            log(f'  SKIP {year}年第{week}週: CSV解析失敗')
            continue

        log(f'  疾病数: {len(parsed["diseases"])}, 都道府県数: {len(parsed["prefectures"])}')

        # 87疾病の Excel を更新
        updated = 0
        skipped = 0
        errors  = 0

        for disease in parsed['diseases']:
            dname  = disease['name']
            folder = DISEASES_DIR / safe_dirname(dname)
            xlsx   = folder / f'{safe_dirname(dname)}.xlsx'

            ok = append_week_to_excel(xlsx, disease, parsed)
            if ok:
                updated += 1
            else:
                skipped += 1

        log(f'  Excel更新: {updated}件成功 / {skipped}件スキップ / {errors}件エラー')

        # 生CSVを保存（バックアップ）
        backup_dir = BASE_DIR / 'idwr_raw' / str(year)
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f'{year}-{week:02d}-zensu.csv'
        with open(backup_path, 'wb') as f:
            f.write(csv_bytes)
        log(f'  CSVバックアップ保存: {backup_path}')

        # 状態を更新
        state['processed_weeks'].append(report['key'])
        state['last_run'] = datetime.now().isoformat()
        state[f'{report["key"]}_updated'] = updated
        save_state(state)

        results.append({
            'week': f'{year}年第{week}週',
            'updated': updated,
            'skipped': skipped
        })

    # 最終サマリー
    log('\n' + '=' * 60)
    log('実行サマリー')
    for r in results:
        log(f'  {r["week"]}: Excel {r["updated"]}件更新')
    log('スクリプト終了')
    log('=' * 60)

if __name__ == '__main__':
    main()
